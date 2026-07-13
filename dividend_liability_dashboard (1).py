"""
Dividend Liability Dashboard
=============================
Tracks the portfolio's cumulative dividend yield (its dividend "liability"
expressed in yield terms) and each holding's contribution to it.

WHY "PORTFOLIO WEIGHT" INSTEAD OF A DOLLAR MARKET VALUE
---------------------------------------------------------
You enter each holding's weight in the portfolio (%) directly, rather than
a dollar market value. Because of that, "dividend liability" here is
expressed in yield/percentage-point terms rather than dollars: each
holding's weight, multiplied by its yield, tells you how many percentage
points of the portfolio's overall dividend yield that holding is
responsible for. Sum those contributions across all holdings and you get
the portfolio's cumulative dividend yield -- the portfolio-level "dividend
liability" this dashboard is built around.

HOW TO RUN
----------
1. Install dependencies:
     pip install matplotlib openpyxl yfinance requests beautifulsoup4
   (tkinter ships with standard Python on Windows/macOS; on Linux you may
    need: sudo apt-get install python3-tk)
   Only matplotlib and openpyxl are required for charts/Excel import.
   yfinance/requests/beautifulsoup4 are optional -- without them, Excel
   import still works, it just won't auto-fetch yields from the web.
2. Open this file in PyCharm and click Run.

IMPORTING A PORTFOLIO FROM EXCEL
---------------------------------
Click "Import Excel". The file needs a header row with at least a Ticker
(or Symbol) column and a Weight (or Weight %) column. TTM Yield / SEC
Yield columns are optional -- if a holding is missing either yield, the
dashboard automatically looks it up on the web (see below) as soon as the
import finishes.

You can also click "Refresh Yields from Web" at any time to re-fetch yields
for every holding currently on the table (e.g. to update stale numbers).

WHERE THE YIELD DATA COMES FROM
---------------------------------
Yield lookups try, in order:
  1. yfinance (pulls Yahoo Finance's published figures) -- gives a TTM
     ("trailing annual dividend yield") figure for stocks, and a fund
     "yield" figure for ETFs/mutual funds that is often the closest public
     proxy for a 30-day SEC yield.
  2. A best-effort HTML scrape of stockanalysis.com as a fallback if
     yfinance is unavailable or doesn't have the ticker, which only
     supplies a TTM-style figure.
These are public, unofficial data sources -- they can lag, be missing for
some tickers, or occasionally be wrong. Always sanity-check the fetched
yields (especially 30-day SEC yield, which for bond/income funds is best
confirmed against the fund issuer's own factsheet) before relying on the
liability numbers for anything important.

READING THE DASHBOARD
----------------------
The Yield Source column (and the color of each row) shows where that
holding's yields came from: blue = Yahoo Finance via yfinance, gold =
the stockanalysis.com fallback scrape, gray = typed in manually or loaded
from a file, red = a fetch was attempted and failed. Hover the legend
under the table for the same key. The right-hand panel has stat cards
(cumulative yields, spread, concentration, web-sourced coverage, last
refresh time) plus two chart tabs: an allocation pie and a top-contributors
bar chart comparing each holding's TTM vs. SEC contribution.

DEFINITIONS USED
-----------------
- Portfolio Weight (%)            = what you enter for each holding; should
                                     sum to ~100% across all holdings.
- Contribution to TTM Yield (pp)  = Weight% x TTM Yield% / 100
- Contribution to SEC Yield (pp)  = Weight% x 30-Day SEC Yield% / 100
    (SEC yield is already an SEC-standardized ANNUALIZED figure, so no
     extra annualization is applied -- it's used as-is.)
- Cumulative Portfolio TTM Yield  = sum of all "Contribution to TTM Yield"
                                     values -> the portfolio's trailing
                                     dividend liability, in yield terms.
- Cumulative Portfolio SEC Yield  = sum of all "Contribution to SEC Yield"
                                     values -> the portfolio's forward-
                                     looking, regulator-standardized
                                     dividend liability.
- Overall Contribution %          = each holding's average of its TTM and
                                     SEC contributions, as a % share of the
                                     portfolio-wide total -- i.e. "how much
                                     of the firm's total dividend liability
                                     does this holding account for."
- Yield Spread (bps)              = (TTM Yield - SEC Yield) x 100
    A large positive spread can indicate a recent yield cut or declining
    distributions (TTM still reflects the older, higher payouts); a large
    negative spread can indicate a recent distribution increase.
- HHI (concentration)             = sum(Overall Contribution%^2), 0-10,000
    scale (10,000 = a single holding accounts for the entire liability;
    lower is more diversified.)
"""

import csv
import queue
import re
import sys
import threading
import time
import tkinter as tk
from dataclasses import dataclass
from datetime import datetime
from tkinter import ttk, messagebox, filedialog
from typing import List, Optional, Tuple

try:
    from matplotlib.figure import Figure
    from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
    MATPLOTLIB_AVAILABLE = True
except ImportError:
    MATPLOTLIB_AVAILABLE = False

try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    import yfinance as yf
    YFINANCE_AVAILABLE = True
except ImportError:
    YFINANCE_AVAILABLE = False

try:
    import requests
    from bs4 import BeautifulSoup
    SCRAPER_AVAILABLE = True
except ImportError:
    SCRAPER_AVAILABLE = False


# --------------------------------------------------------------------------
# Web yield lookup
# --------------------------------------------------------------------------

def fetch_yield_data(ticker: str) -> Tuple[Optional[float], Optional[float], str]:
    """
    Best-effort lookup of a ticker's TTM dividend yield and 30-day SEC yield.

    Tries yfinance first (Yahoo Finance's published figures), then falls
    back to scraping stockanalysis.com if yfinance is unavailable or has no
    data for the ticker. Either returned yield may be None if it couldn't be
    found anywhere.

    Unlike an earlier version of this function, errors are NOT swallowed --
    if a source fails, a short reason ("rate-limited", "no internet", the
    exception text, etc.) is folded into the returned source_description so
    it's visible in the dashboard's status bar and in the console, instead of
    silently doing nothing.

    Returns: (ttm_yield_pct, sec_yield_pct, source_description)
    """
    ttm: Optional[float] = None
    sec: Optional[float] = None
    notes = []

    if YFINANCE_AVAILABLE:
        yf_error = None
        # Yahoo aggressively rate-limits yfinance (HTTP 429 / YFRateLimitError)
        # when tickers are queried back-to-back, so retry once after a short
        # pause before giving up on this source.
        for attempt in range(2):
            try:
                info = yf.Ticker(ticker).info
                trailing = info.get("trailingAnnualDividendYield")
                if trailing is not None:
                    ttm = float(trailing) * 100.0
                # For ETFs/mutual funds, Yahoo's "yield" field is the closest
                # public proxy to a 30-day SEC yield.
                fund_yield = info.get("yield")
                if fund_yield is not None:
                    sec = float(fund_yield) * 100.0
                if ttm is not None or sec is not None:
                    notes.append("Yahoo Finance (yfinance)")
                else:
                    yf_error = "yfinance returned no dividend/yield fields for this ticker"
                yf_error = None
                break
            except Exception as exc:
                yf_error = f"{type(exc).__name__}: {exc}"
                if "rate" in yf_error.lower() or "429" in yf_error:
                    time.sleep(2.0)
                    continue
                break
        if yf_error:
            print(f"[fetch_yield_data] yfinance failed for {ticker}: {yf_error}", file=sys.stderr)
            notes.append(f"yfinance error ({yf_error})")

    if (ttm is None or sec is None) and SCRAPER_AVAILABLE:
        try:
            scraped_ttm = _scrape_stockanalysis_ttm_yield(ticker)
            if ttm is None and scraped_ttm is not None:
                ttm = scraped_ttm
                notes.append("stockanalysis.com")
            elif ttm is None:
                notes.append("stockanalysis.com had no match")
        except Exception as exc:
            err = f"{type(exc).__name__}: {exc}"
            print(f"[fetch_yield_data] stockanalysis.com scrape failed for {ticker}: {err}", file=sys.stderr)
            notes.append(f"scrape error ({err})")

    if not YFINANCE_AVAILABLE and not SCRAPER_AVAILABLE:
        notes.append("no fetch library installed")

    # Being polite to Yahoo/stockanalysis.com reduces the odds of the next
    # ticker in the batch getting rate-limited.
    time.sleep(0.6)

    return ttm, sec, "; ".join(notes) if notes else "no source found"


def _scrape_stockanalysis_ttm_yield(ticker: str) -> Optional[float]:
    """
    Fallback scraper: pulls a TTM-style dividend yield figure off a
    stockanalysis.com stock or ETF page. This depends on that site's current
    page wording/layout and is meant purely as a fallback when yfinance has
    no data -- treat it as approximate, not authoritative. It does not
    supply a 30-day SEC yield (no single public site standardizes that
    figure across tickers; for funds it's best confirmed on the issuer's
    own factsheet).
    """
    headers = {"User-Agent": "Mozilla/5.0 (dividend-liability-dashboard)"}
    for url in (
        f"https://stockanalysis.com/stocks/{ticker.lower()}/",
        f"https://stockanalysis.com/etf/{ticker.lower()}/",
    ):
        try:
            resp = requests.get(url, headers=headers, timeout=8)
            if resp.status_code != 200:
                continue
            text = BeautifulSoup(resp.text, "html.parser").get_text(" ", strip=True)
            match = re.search(r"Dividend Yield[^0-9%]{0,20}([\d.]+)\s*%", text, re.IGNORECASE)
            if match:
                return float(match.group(1))
        except Exception:
            continue
    return None


# --------------------------------------------------------------------------
# Data model
# --------------------------------------------------------------------------

@dataclass
class Holding:
    ticker: str
    weight_pct: float      # e.g. 12.5 means 12.5% of the portfolio
    ttm_yield_pct: float   # e.g. 3.50 means 3.50%
    sec_yield_pct: float   # e.g. 3.20 means 3.20%
    yield_source: str = "Manual entry"   # where the yield figures came from
    last_fetched: str = ""               # timestamp of last web fetch, "" if never

    @property
    def ttm_contribution_pp(self) -> float:
        """Contribution to cumulative portfolio TTM yield, in percentage points."""
        return self.weight_pct / 100.0 * self.ttm_yield_pct

    @property
    def sec_contribution_pp(self) -> float:
        """Contribution to cumulative portfolio SEC yield, in percentage points."""
        return self.weight_pct / 100.0 * self.sec_yield_pct

    @property
    def avg_contribution_pp(self) -> float:
        return (self.ttm_contribution_pp + self.sec_contribution_pp) / 2.0

    @property
    def yield_spread_bps(self) -> float:
        return (self.ttm_yield_pct - self.sec_yield_pct) * 100.0


def classify_source(raw_source: str) -> Tuple[str, str]:
    """
    Turns a raw source string (either a fixed label like 'Manual entry', or
    the note string returned by fetch_yield_data) into a short badge label
    and a Treeview tag name used to color that row by provenance.
    """
    if raw_source in ("Manual entry", "CSV import", "Excel file"):
        return raw_source, "src_manual"
    if "Yahoo Finance" in raw_source:
        return "Yahoo Finance", "src_web_primary"
    if "stockanalysis.com" in raw_source and "no match" not in raw_source and "error" not in raw_source:
        return "stockanalysis.com (fallback)", "src_web_fallback"
    if "no fetch library" in raw_source:
        return "No fetch library installed", "src_failed"
    return "Fetch failed", "src_failed"


# --------------------------------------------------------------------------
# Main application
# --------------------------------------------------------------------------

class DividendDashboard(tk.Tk):

    # ---- Design tokens ---------------------------------------------------
    # A "financial terminal" palette: deep slate background (not flat black),
    # a warm gold accent that nods to the dividend/income subject matter
    # rather than a generic neon-on-black look, and tabular monospace
    # numerals in the stat cards for a data-terminal feel.
    BG_APP = "#141922"
    BG_PANEL = "#1B222D"
    BG_PANEL_ALT = "#212A37"
    BG_INPUT = "#242E3B"
    BORDER = "#2E3A48"
    TEXT_PRIMARY = "#EAEDF2"
    TEXT_SECONDARY = "#8B96A6"
    TEXT_MUTED = "#5D6B7D"
    ACCENT = "#E3AB4E"          # warm gold -- dividend/income signature color
    ACCENT_DIM = "#8A6A34"
    POSITIVE = "#4FBF8B"        # muted emerald
    NEGATIVE = "#E2685C"        # muted coral
    INFO = "#5CA3D9"            # muted blue, used for web-fetched badges

    FONT_UI = ("Segoe UI", 10)
    FONT_UI_BOLD = ("Segoe UI", 10, "bold")
    FONT_HEADER = ("Segoe UI", 17, "bold")
    FONT_SUB = ("Segoe UI", 9)
    FONT_MONO = ("Consolas", 10)
    FONT_MONO_BIG = ("Consolas", 20, "bold")
    FONT_MONO_SMALL = ("Consolas", 9)

    COLUMNS = [
        ("ticker", "Ticker", 65),
        ("weight", "Wt %", 65),
        ("ttm_yield", "TTM Yield %", 85),
        ("sec_yield", "SEC Yield %", 90),
        ("spread", "Spread (bps)", 90),
        ("ttm_contrib", "TTM Contrib. (pp)", 130),
        ("sec_contrib", "SEC Contrib. (pp)", 130),
        ("overall_contrib", "Overall %", 85),
        ("source", "Yield Source", 230),
    ]

    def __init__(self):
        super().__init__()
        self.title("Dividend Liability Dashboard")
        self.geometry("1780x920")
        self.minsize(1500, 740)
        self.configure(bg=self.BG_APP)

        self._resolve_fonts()

        self.holdings: List[Holding] = []

        self._build_style()
        self._build_input_bar()
        self._build_body()
        self._build_status_bar()

        self._refresh_all()

    # ------------------------------------------------------------ fonts --
    def _resolve_fonts(self):
        """Falls back to fonts that actually exist on this OS (Segoe UI and
        Consolas are Windows-only; macOS/Linux get sensible equivalents)."""
        import tkinter.font as tkfont
        available = set(tkfont.families())

        def pick(preferred, fallbacks):
            for name in [preferred] + fallbacks:
                if name in available:
                    return name
            return preferred

        ui_face = pick("Segoe UI", ["Helvetica Neue", "Helvetica", "Arial", "DejaVu Sans"])
        mono_face = pick("Consolas", ["Menlo", "SF Mono", "DejaVu Sans Mono", "Courier New"])

        self.FONT_UI = (ui_face, 10)
        self.FONT_UI_BOLD = (ui_face, 10, "bold")
        self.FONT_HEADER = (ui_face, 17, "bold")
        self.FONT_SUB = (ui_face, 9)
        self.FONT_MONO = (mono_face, 10)
        self.FONT_MONO_BIG = (mono_face, 20, "bold")
        self.FONT_MONO_SMALL = (mono_face, 9)

    # ---------------------------------------------------------- styling --
    def _build_style(self):
        style = ttk.Style(self)
        try:
            style.theme_use("clam")
        except tk.TclError:
            pass

        style.configure(".", background=self.BG_APP, foreground=self.TEXT_PRIMARY, font=self.FONT_UI)
        style.configure("TFrame", background=self.BG_APP)
        style.configure("Panel.TFrame", background=self.BG_PANEL)
        style.configure("Card.TFrame", background=self.BG_PANEL_ALT)

        style.configure("TLabel", background=self.BG_APP, foreground=self.TEXT_PRIMARY, font=self.FONT_UI)
        style.configure("Panel.TLabel", background=self.BG_PANEL, foreground=self.TEXT_PRIMARY, font=self.FONT_UI)
        style.configure("Card.TLabel", background=self.BG_PANEL_ALT, foreground=self.TEXT_PRIMARY, font=self.FONT_UI)
        style.configure("Header.TLabel", background=self.BG_APP, foreground=self.TEXT_PRIMARY, font=self.FONT_HEADER)
        style.configure("Accent.TLabel", background=self.BG_APP, foreground=self.ACCENT, font=self.FONT_UI_BOLD)
        style.configure("Sub.TLabel", background=self.BG_APP, foreground=self.TEXT_SECONDARY, font=self.FONT_SUB)
        style.configure("Warn.TLabel", background=self.BG_APP, foreground=self.NEGATIVE, font=(self.FONT_UI[0], 9, "bold"))

        style.configure("CardLabel.TLabel", background=self.BG_PANEL_ALT, foreground=self.TEXT_SECONDARY, font=self.FONT_SUB)
        style.configure("CardValue.TLabel", background=self.BG_PANEL_ALT, foreground=self.TEXT_PRIMARY, font=self.FONT_MONO_BIG)
        style.configure("CardValueSmall.TLabel", background=self.BG_PANEL_ALT, foreground=self.TEXT_PRIMARY, font=(self.FONT_MONO[0], 13, "bold"))
        style.configure("CardSub.TLabel", background=self.BG_PANEL_ALT, foreground=self.TEXT_MUTED, font=self.FONT_MONO_SMALL)

        style.configure("TEntry", fieldbackground=self.BG_INPUT, foreground=self.TEXT_PRIMARY,
                         insertcolor=self.TEXT_PRIMARY, bordercolor=self.BORDER, lightcolor=self.BORDER,
                         darkcolor=self.BORDER, padding=5)
        style.map("TEntry", fieldbackground=[("focus", self.BG_INPUT)], bordercolor=[("focus", self.ACCENT)])

        style.configure("TButton", background=self.BG_INPUT, foreground=self.TEXT_PRIMARY,
                         font=self.FONT_UI, padding=(10, 6), borderwidth=0)
        style.map("TButton", background=[("active", self.BG_PANEL_ALT)], foreground=[("active", self.TEXT_PRIMARY)])

        style.configure("Accent.TButton", background=self.ACCENT, foreground="#1B1400",
                         font=self.FONT_UI_BOLD, padding=(12, 6), borderwidth=0)
        style.map("Accent.TButton", background=[("active", "#F0BC66")])

        style.configure("TNotebook", background=self.BG_PANEL, bordercolor=self.BORDER, tabmargins=(0, 4, 0, 0))
        style.configure("TNotebook.Tab", background=self.BG_PANEL, foreground=self.TEXT_SECONDARY,
                         font=self.FONT_SUB, padding=(12, 5), borderwidth=0)
        style.map("TNotebook.Tab",
                  background=[("selected", self.BG_PANEL_ALT)],
                  foreground=[("selected", self.ACCENT)])

        style.configure("TLabelframe", background=self.BG_APP, bordercolor=self.BORDER, darkcolor=self.BG_APP, lightcolor=self.BG_APP)
        style.configure("TLabelframe.Label", background=self.BG_APP, foreground=self.TEXT_SECONDARY, font=self.FONT_UI_BOLD)

        style.configure("Treeview", background=self.BG_PANEL, fieldbackground=self.BG_PANEL,
                         foreground=self.TEXT_PRIMARY, rowheight=27, font=self.FONT_MONO, borderwidth=0)
        style.configure("Treeview.Heading", background=self.BG_PANEL_ALT, foreground=self.TEXT_SECONDARY,
                         font=(self.FONT_UI[0], 9, "bold"), relief="flat", borderwidth=0)
        style.map("Treeview.Heading", background=[("active", self.BG_PANEL_ALT)])
        style.map("Treeview",
                  background=[("selected", self.ACCENT_DIM)],
                  foreground=[("selected", self.TEXT_PRIMARY)])

        style.configure("Vertical.TScrollbar", background=self.BG_PANEL_ALT, troughcolor=self.BG_APP,
                         bordercolor=self.BG_APP, arrowcolor=self.TEXT_SECONDARY)

    # ------------------------------------------------------- input bar --
    def _build_input_bar(self):
        outer = tk.Frame(self, bg=self.BG_APP, padx=16, pady=14)
        outer.pack(side="top", fill="x")

        title_row = tk.Frame(outer, bg=self.BG_APP)
        title_row.pack(side="top", fill="x", pady=(0, 12))
        ttk.Label(title_row, text="Dividend Liability Dashboard", style="Header.TLabel").pack(side="left")
        ttk.Label(
            title_row, text="  ·  portfolio yield exposure & data provenance", style="Sub.TLabel"
        ).pack(side="left", padx=(2, 0))

        bar = tk.Frame(outer, bg=self.BG_PANEL, padx=14, pady=12, highlightbackground=self.BORDER,
                        highlightthickness=1)
        bar.pack(side="top", fill="x")

        fields = [
            ("Ticker", "ticker_var", 8),
            ("Portfolio Weight (%)", "weight_var", 14),
            ("TTM Yield (%)", "ttm_var", 12),
            ("30D SEC Yield (%)", "sec_var", 14),
        ]
        self.entry_vars = {}
        col = 0
        for label_text, var_name, width in fields:
            tk.Label(bar, text=label_text, bg=self.BG_PANEL, fg=self.TEXT_SECONDARY, font=self.FONT_SUB).grid(
                row=0, column=col, sticky="w", padx=(0, 4)
            )
            var = tk.StringVar()
            self.entry_vars[var_name] = var
            entry = ttk.Entry(bar, textvariable=var, width=width, font=self.FONT_UI)
            entry.grid(row=1, column=col, sticky="w", padx=(0, 16), ipady=2)
            col += 1

        add_btn = ttk.Button(bar, text="+  Add Holding", command=self._add_holding, style="Accent.TButton")
        add_btn.grid(row=1, column=col, sticky="w", padx=(4, 20))
        col += 1

        sep = tk.Frame(bar, bg=self.BORDER, width=1)
        sep.grid(row=0, column=col, rowspan=2, sticky="ns", padx=(0, 16))
        col += 1

        button_defs = [
            ("Remove Selected", self._remove_selected),
            ("Clear All", self._clear_all),
            ("Export CSV", self._export_csv),
            ("Import CSV", self._import_csv),
            ("Import Excel", self._import_excel),
        ]
        for text, cmd in button_defs:
            ttk.Button(bar, text=text, command=cmd).grid(row=1, column=col, sticky="w", padx=(0, 8))
            col += 1

        refresh_btn = ttk.Button(
            bar, text="\u21bb  Refresh Yields", command=self._refresh_yields_from_web,
            style="Accent.TButton",
        )
        refresh_btn.grid(row=1, column=col, sticky="w", padx=(8, 0))
        col += 1

        ttk.Label(
            outer,
            text=("Enter weight and yields as plain numbers, e.g. 12.5 for 12.5%. Weights should sum to ~100%. "
                  "Press Enter to add quickly. Yields fetched from the web are labeled in the Yield Source column."),
            style="Sub.TLabel",
        ).pack(side="top", anchor="w", pady=(8, 0))

        self.bind_all("<Return>", lambda e: self._add_holding())

    # ----------------------------------------------------------- body --
    def _build_body(self):
        body = tk.Frame(self, bg=self.BG_APP, padx=16, pady=6)
        body.pack(side="top", fill="both", expand=True)

        # ---- Left: table ----
        left = tk.Frame(body, bg=self.BG_APP)
        left.pack(side="left", fill="both", expand=True)

        table_card = tk.Frame(left, bg=self.BG_PANEL, highlightbackground=self.BORDER, highlightthickness=1)
        table_card.pack(side="top", fill="both", expand=True)

        columns = [c[0] for c in self.COLUMNS]
        self.tree = ttk.Treeview(table_card, columns=columns, show="headings", selectmode="browse")
        for key, heading, width in self.COLUMNS:
            self.tree.heading(key, text=heading)
            anchor = "w" if key == "source" else "center"
            self.tree.column(key, width=width, anchor=anchor)
        self.tree.pack(side="left", fill="both", expand=True, padx=(1, 0), pady=1)

        # Row tags: alternating banding + one color per yield-source provenance
        self.tree.tag_configure("even", background=self.BG_PANEL)
        self.tree.tag_configure("odd", background=self.BG_PANEL_ALT)
        self.tree.tag_configure("src_web_primary", foreground=self.INFO)
        self.tree.tag_configure("src_web_fallback", foreground=self.ACCENT)
        self.tree.tag_configure("src_manual", foreground=self.TEXT_SECONDARY)
        self.tree.tag_configure("src_failed", foreground=self.NEGATIVE)

        vsb = ttk.Scrollbar(table_card, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y")

        legend = tk.Frame(left, bg=self.BG_APP)
        legend.pack(side="top", fill="x", pady=(6, 0))
        for text, color in [
            ("\u25cf Yahoo Finance (web)", self.INFO),
            ("\u25cf stockanalysis.com (fallback)", self.ACCENT),
            ("\u25cf Manual / file entry", self.TEXT_SECONDARY),
            ("\u25cf Fetch failed", self.NEGATIVE),
        ]:
            tk.Label(legend, text=text, bg=self.BG_APP, fg=color, font=self.FONT_SUB).pack(side="left", padx=(0, 16))

        # ---- Right: stat cards + charts ----
        right = tk.Frame(body, width=420, bg=self.BG_APP)
        right.pack(side="right", fill="y", padx=(14, 0))
        right.pack_propagate(False)

        tk.Label(right, text="PORTFOLIO STATISTICS", bg=self.BG_APP, fg=self.TEXT_SECONDARY,
                 font=(self.FONT_UI[0], 9, "bold")).pack(side="top", anchor="w", pady=(0, 6))

        cards_grid = tk.Frame(right, bg=self.BG_APP)
        cards_grid.pack(side="top", fill="x")
        cards_grid.columnconfigure(0, weight=1)
        cards_grid.columnconfigure(1, weight=1)

        stat_defs = [
            ("cum_ttm_yield", "Cumulative TTM Yield", "big"),
            ("cum_sec_yield", "Cumulative SEC Yield", "big"),
            ("total_weight", "Total Weight Entered", "small"),
            ("num_holdings", "Holdings", "small"),
            ("avg_spread", "Avg Yield Spread", "small"),
            ("hhi", "Concentration (HHI)", "small"),
            ("web_coverage", "Web-Sourced Yields", "small"),
            ("last_refreshed", "Last Refreshed", "small"),
            ("top_contributor", "Top Contributor", "small"),
            ("avg_yield_pair", "Avg TTM / SEC Yield", "small"),
        ]
        self.stat_cards = {}
        for i, (key, label, size) in enumerate(stat_defs):
            r, c = divmod(i, 2)
            card = tk.Frame(cards_grid, bg=self.BG_PANEL_ALT, highlightbackground=self.BORDER,
                             highlightthickness=1, padx=10, pady=8)
            card.grid(row=r, column=c, sticky="nsew", padx=(0 if c == 0 else 6, 6 if c == 0 else 0), pady=6)

            accent_bar = tk.Frame(card, bg=self.ACCENT, width=3)
            accent_bar.pack(side="left", fill="y", padx=(0, 8))

            text_col = tk.Frame(card, bg=self.BG_PANEL_ALT)
            text_col.pack(side="left", fill="both", expand=True)

            ttk.Label(text_col, text=label.upper(), style="CardLabel.TLabel").pack(side="top", anchor="w")
            value_style = "CardValue.TLabel" if size == "big" else "CardValueSmall.TLabel"
            val = ttk.Label(text_col, text="--", style=value_style)
            val.pack(side="top", anchor="w", pady=(2, 0))
            self.stat_cards[key] = val

        self.weight_warning = ttk.Label(right, text="", style="Warn.TLabel", wraplength=400, justify="left")
        self.weight_warning.pack(side="top", anchor="w", pady=(4, 6))

        # ---- Charts, tabbed ----
        chart_card = tk.Frame(right, bg=self.BG_PANEL, highlightbackground=self.BORDER, highlightthickness=1)
        chart_card.pack(side="top", fill="both", expand=True, pady=(4, 0))

        if MATPLOTLIB_AVAILABLE:
            notebook = ttk.Notebook(chart_card)
            notebook.pack(fill="both", expand=True, padx=4, pady=4)

            pie_tab = tk.Frame(notebook, bg=self.BG_PANEL)
            bar_tab = tk.Frame(notebook, bg=self.BG_PANEL)
            notebook.add(pie_tab, text="Allocation")
            notebook.add(bar_tab, text="Top Contributors")

            self._mpl_dark = {
                "figure.facecolor": self.BG_PANEL,
                "axes.facecolor": self.BG_PANEL,
                "axes.edgecolor": self.BORDER,
                "axes.labelcolor": self.TEXT_SECONDARY,
                "text.color": self.TEXT_PRIMARY,
                "xtick.color": self.TEXT_SECONDARY,
                "ytick.color": self.TEXT_SECONDARY,
            }

            self.fig = Figure(figsize=(3.8, 3.4), dpi=100)
            self.fig.patch.set_facecolor(self.BG_PANEL)
            self.ax = self.fig.add_subplot(111)
            self.canvas = FigureCanvasTkAgg(self.fig, master=pie_tab)
            self.canvas.get_tk_widget().pack(fill="both", expand=True)

            self.fig_bar = Figure(figsize=(3.8, 3.4), dpi=100)
            self.fig_bar.patch.set_facecolor(self.BG_PANEL)
            self.ax_bar = self.fig_bar.add_subplot(111)
            self.canvas_bar = FigureCanvasTkAgg(self.fig_bar, master=bar_tab)
            self.canvas_bar.get_tk_widget().pack(fill="both", expand=True)
        else:
            ttk.Label(
                chart_card,
                text="matplotlib not installed.\nRun: pip install matplotlib\nto enable charts.",
                style="Panel.TLabel",
                justify="center",
            ).pack(expand=True)

    # ------------------------------------------------------ status bar --
    def _build_status_bar(self):
        self.status_var = tk.StringVar(value="Ready.")
        bar = tk.Frame(self, bg=self.BG_PANEL, padx=16, pady=6, highlightbackground=self.BORDER, highlightthickness=1)
        bar.pack(side="bottom", fill="x")
        tk.Label(bar, text="\u25cf", bg=self.BG_PANEL, fg=self.POSITIVE, font=self.FONT_SUB).pack(side="left", padx=(0, 6))
        ttk.Label(bar, textvariable=self.status_var, style="Panel.TLabel").pack(side="left")

    # ------------------------------------------------------------ logic --
    def _add_holding(self):
        ticker = self.entry_vars["ticker_var"].get().strip().upper()
        weight_raw = self.entry_vars["weight_var"].get().strip()
        ttm_raw = self.entry_vars["ttm_var"].get().strip()
        sec_raw = self.entry_vars["sec_var"].get().strip()

        if not ticker or not weight_raw or not ttm_raw or not sec_raw:
            return  # silently ignore empty submissions (e.g. stray Enter presses)

        try:
            weight = float(weight_raw.replace("%", ""))
            ttm_yield = float(ttm_raw.replace("%", ""))
            sec_yield = float(sec_raw.replace("%", ""))
        except ValueError:
            messagebox.showerror(
                "Invalid input",
                "Portfolio Weight, TTM Yield, and SEC Yield must be numbers.\n"
                "Example: Portfolio Weight = 12.5, TTM Yield = 3.5"
            )
            return

        if weight <= 0:
            messagebox.showerror("Invalid input", "Portfolio Weight must be greater than 0.")
            return
        if ttm_yield < 0 or sec_yield < 0:
            messagebox.showerror("Invalid input", "Yields cannot be negative.")
            return

        # If ticker already exists, update it instead of duplicating
        existing = next((h for h in self.holdings if h.ticker == ticker), None)
        if existing:
            if not messagebox.askyesno(
                "Ticker exists",
                f"{ticker} is already in the portfolio. Update it with the new values?"
            ):
                return
            existing.weight_pct = weight
            existing.ttm_yield_pct = ttm_yield
            existing.sec_yield_pct = sec_yield
            existing.yield_source = "Manual entry"
            existing.last_fetched = ""
        else:
            self.holdings.append(Holding(ticker, weight, ttm_yield, sec_yield, "Manual entry", ""))

        for var_name in ("ticker_var", "weight_var", "ttm_var", "sec_var"):
            self.entry_vars[var_name].set("")

        self._refresh_all()
        self.status_var.set(f"Added/updated {ticker}.")

    def _remove_selected(self):
        selection = self.tree.selection()
        if not selection:
            messagebox.showinfo("Nothing selected", "Select a row in the table first.")
            return
        ticker = self.tree.item(selection[0], "values")[0]
        self.holdings = [h for h in self.holdings if h.ticker != ticker]
        self._refresh_all()
        self.status_var.set(f"Removed {ticker}.")

    def _clear_all(self):
        if not self.holdings:
            return
        if messagebox.askyesno("Clear all", "Remove all holdings from the dashboard?"):
            self.holdings.clear()
            self._refresh_all()
            self.status_var.set("Cleared all holdings.")

    def _export_csv(self):
        if not self.holdings:
            messagebox.showinfo("No data", "Add holdings before exporting.")
            return
        path = filedialog.asksaveasfilename(
            defaultextension=".csv",
            filetypes=[("CSV files", "*.csv")],
            title="Export holdings to CSV",
        )
        if not path:
            return
        with open(path, "w", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(["Ticker", "WeightPct", "TTMYieldPct", "SECYieldPct", "YieldSource", "LastFetched"])
            for h in self.holdings:
                writer.writerow([h.ticker, h.weight_pct, h.ttm_yield_pct, h.sec_yield_pct, h.yield_source, h.last_fetched])
        self.status_var.set(f"Exported {len(self.holdings)} holdings to {path}")

    def _import_csv(self):
        path = filedialog.askopenfilename(filetypes=[("CSV files", "*.csv")], title="Import holdings CSV")
        if not path:
            return
        try:
            with open(path, newline="") as f:
                reader = csv.DictReader(f)
                imported = 0
                for row in reader:
                    ticker = row["Ticker"].strip().upper()
                    weight = float(row["WeightPct"])
                    ttm = float(row["TTMYieldPct"])
                    sec = float(row["SECYieldPct"])
                    source = row.get("YieldSource", "").strip() or "CSV import"
                    last_fetched = row.get("LastFetched", "").strip()
                    existing = next((h for h in self.holdings if h.ticker == ticker), None)
                    if existing:
                        existing.weight_pct, existing.ttm_yield_pct, existing.sec_yield_pct = weight, ttm, sec
                        existing.yield_source, existing.last_fetched = source, last_fetched
                    else:
                        self.holdings.append(Holding(ticker, weight, ttm, sec, source, last_fetched))
                    imported += 1
            self._refresh_all()
            self.status_var.set(f"Imported {imported} holdings from {path}")
        except Exception as exc:
            messagebox.showerror("Import failed", f"Could not import file:\n{exc}")

    # ------------------------------------------------------ excel import --
    def _import_excel(self):
        if not OPENPYXL_AVAILABLE:
            messagebox.showerror(
                "openpyxl not installed",
                "Reading Excel files requires openpyxl.\n\nRun:\n  pip install openpyxl"
            )
            return

        path = filedialog.askopenfilename(
            filetypes=[("Excel files", "*.xlsx *.xlsm")],
            title="Import portfolio from Excel",
        )
        if not path:
            return

        try:
            rows = self._read_excel_rows(path)
        except Exception as exc:
            messagebox.showerror("Import failed", f"Could not read Excel file:\n{exc}")
            return

        if not rows:
            messagebox.showinfo("No data", "No holdings found in that Excel file.")
            return

        tickers_needing_yields = []
        for ticker, weight, ttm, sec in rows:
            existing = next((h for h in self.holdings if h.ticker == ticker), None)
            if existing:
                existing.weight_pct = weight
                if ttm is not None:
                    existing.ttm_yield_pct = ttm
                if sec is not None:
                    existing.sec_yield_pct = sec
                if ttm is not None and sec is not None:
                    existing.yield_source = "Excel file"
                    existing.last_fetched = ""
            else:
                source = "Excel file" if (ttm is not None and sec is not None) else "Manual entry"
                self.holdings.append(Holding(ticker, weight, ttm or 0.0, sec or 0.0, source, ""))
            if ttm is None or sec is None:
                tickers_needing_yields.append(ticker)

        self._refresh_all()
        self.status_var.set(f"Imported {len(rows)} holdings from Excel.")

        if tickers_needing_yields:
            if not (YFINANCE_AVAILABLE or SCRAPER_AVAILABLE):
                messagebox.showwarning(
                    "Yield lookup unavailable",
                    "Your Excel file didn't include TTM/SEC yields for every holding, "
                    "and no yield-fetching library is installed.\n\nRun:\n"
                    "  pip install yfinance requests beautifulsoup4\n\n"
                    "then click 'Refresh Yields from Web'."
                )
                return
            self._fetch_yields_for_tickers(tickers_needing_yields)

    def _read_excel_rows(self, path):
        """
        Reads (ticker, weight_pct, ttm_yield_pct_or_None, sec_yield_pct_or_None)
        tuples from an Excel file's first sheet. Expects a header row with a
        Ticker/Symbol column and a Weight/Weight % column; TTM Yield and SEC
        Yield columns are optional.
        """
        wb = load_workbook(path, data_only=True, read_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)

        try:
            header = next(rows_iter)
        except StopIteration:
            return []

        header_map = {}
        for idx, h in enumerate(header):
            key = self._normalize_header(h)
            if key:
                header_map[key] = idx

        ticker_col = self._first_matching_col(header_map, ["ticker", "symbol"])
        weight_col = self._first_matching_col(
            header_map, ["weight", "weightpct", "portfolioweight", "portfolioweightpct"]
        )
        ttm_col = self._first_matching_col(header_map, ["ttmyield", "ttmyieldpct", "ttm"])
        sec_col = self._first_matching_col(
            header_map, ["secyield", "secyieldpct", "30dsecyield", "30daysecyield", "sec"]
        )

        if ticker_col is None or weight_col is None:
            raise ValueError(
                "Couldn't find 'Ticker' and 'Weight' columns. Expected headers like "
                "'Ticker' (or 'Symbol') and 'Weight' (or 'Weight %'). "
                "'TTM Yield' / 'SEC Yield' columns are optional."
            )

        results = []
        for row in rows_iter:
            if row is None or all(v is None for v in row):
                continue
            ticker = row[ticker_col] if ticker_col < len(row) else None
            weight = row[weight_col] if weight_col < len(row) else None
            if ticker is None or weight is None or str(ticker).strip() == "":
                continue

            ticker = str(ticker).strip().upper()
            weight = float(str(weight).replace("%", ""))

            ttm = None
            if ttm_col is not None and ttm_col < len(row) and row[ttm_col] is not None:
                ttm = float(str(row[ttm_col]).replace("%", ""))

            sec = None
            if sec_col is not None and sec_col < len(row) and row[sec_col] is not None:
                sec = float(str(row[sec_col]).replace("%", ""))

            results.append((ticker, weight, ttm, sec))

        return results

    @staticmethod
    def _normalize_header(h) -> str:
        if h is None:
            return ""
        return re.sub(r"[^a-z0-9]", "", str(h).strip().lower())

    @staticmethod
    def _first_matching_col(header_map, candidates):
        for c in candidates:
            if c in header_map:
                return header_map[c]
        return None

    # -------------------------------------------------- web yield fetch --
    def _refresh_yields_from_web(self):
        if not self.holdings:
            messagebox.showinfo("No holdings", "Add or import holdings first.")
            return
        if not (YFINANCE_AVAILABLE or SCRAPER_AVAILABLE):
            messagebox.showwarning(
                "Yield lookup unavailable",
                "No yield-fetching library is installed.\n\nRun:\n"
                "  pip install yfinance requests beautifulsoup4"
            )
            return
        tickers = [h.ticker for h in self.holdings]
        self._fetch_yields_for_tickers(tickers)

    def _fetch_yields_for_tickers(self, tickers: List[str]):
        """Fetches yields for the given tickers on a background thread so the
        UI doesn't freeze, then applies results as they arrive."""
        result_queue: "queue.Queue" = queue.Queue()
        total = len(tickers)
        self.status_var.set(f"Fetching yields from the web for {total} ticker(s)...")

        def worker():
            for ticker in tickers:
                ttm, sec, source = fetch_yield_data(ticker)
                result_queue.put((ticker, ttm, sec, source))

        threading.Thread(target=worker, daemon=True).start()
        self._poll_yield_results(result_queue, total, {"done": 0})

    def _poll_yield_results(self, result_queue: "queue.Queue", total: int, progress: dict):
        updated_any = False
        try:
            while True:
                ticker, ttm, sec, source = result_queue.get_nowait()
                progress["done"] += 1
                holding = next((h for h in self.holdings if h.ticker == ticker), None)
                badge, _tag = classify_source(source)
                if holding:
                    if ttm is not None:
                        holding.ttm_yield_pct = ttm
                    if sec is not None:
                        holding.sec_yield_pct = sec
                    if ttm is not None or sec is not None:
                        holding.yield_source = badge
                        holding.last_fetched = datetime.now().strftime("%H:%M:%S")
                        updated_any = True
                    else:
                        holding.yield_source = badge  # e.g. "Fetch failed"
                        updated_any = True
                if ttm is None and sec is None:
                    self.status_var.set(f"No yield data found for {ticker} ({progress['done']}/{total}): {source}")
                else:
                    self.status_var.set(
                        f"Fetched {ticker} from {badge} ({progress['done']}/{total})."
                    )
        except queue.Empty:
            pass

        if updated_any:
            self._refresh_all()

        if progress["done"] < total:
            self.after(200, lambda: self._poll_yield_results(result_queue, total, progress))
        else:
            self.status_var.set(f"Finished fetching yields for {total} ticker(s).")
            self._refresh_all()

    # -------------------------------------------------------- refresh --
    def _refresh_all(self):
        self._refresh_table()
        self._refresh_stats()
        self._refresh_pie_chart()
        self._refresh_bar_chart()

    def _refresh_table(self):
        self.tree.delete(*self.tree.get_children())
        total_avg_contrib = sum(h.avg_contribution_pp for h in self.holdings) or 1.0

        # Sort by average contribution descending -> biggest liability drivers first
        for i, h in enumerate(sorted(self.holdings, key=lambda x: x.avg_contribution_pp, reverse=True)):
            overall_contribution_pct = h.avg_contribution_pp / total_avg_contrib * 100.0
            badge, source_tag = classify_source(h.yield_source)
            short_badge = badge.split(" (")[0]
            source_display = short_badge if not h.last_fetched else f"{short_badge} @ {h.last_fetched}"
            band_tag = "even" if i % 2 == 0 else "odd"
            self.tree.insert(
                "", "end",
                values=(
                    h.ticker,
                    f"{h.weight_pct:,.2f}",
                    f"{h.ttm_yield_pct:,.2f}",
                    f"{h.sec_yield_pct:,.2f}",
                    f"{h.yield_spread_bps:,.1f}",
                    f"{h.ttm_contribution_pp:,.3f}",
                    f"{h.sec_contribution_pp:,.3f}",
                    f"{overall_contribution_pct:,.2f}",
                    source_display,
                ),
                tags=(band_tag, source_tag),
            )

    def _refresh_stats(self):
        n = len(self.holdings)
        total_weight = sum(h.weight_pct for h in self.holdings)
        cum_ttm_yield = sum(h.ttm_contribution_pp for h in self.holdings)
        cum_sec_yield = sum(h.sec_contribution_pp for h in self.holdings)
        avg_spread = (sum(h.yield_spread_bps for h in self.holdings) / n) if n else 0.0
        avg_ttm = (sum(h.ttm_yield_pct for h in self.holdings) / n) if n else 0.0
        avg_sec = (sum(h.sec_yield_pct for h in self.holdings) / n) if n else 0.0

        total_avg_contrib = sum(h.avg_contribution_pp for h in self.holdings)
        if total_avg_contrib > 0:
            hhi = sum((h.avg_contribution_pp / total_avg_contrib * 100.0) ** 2 for h in self.holdings)
        else:
            hhi = 0.0

        top_contributor = "--"
        if self.holdings:
            top = max(self.holdings, key=lambda x: x.avg_contribution_pp)
            pct = (top.avg_contribution_pp / total_avg_contrib * 100.0) if total_avg_contrib else 0.0
            top_contributor = f"{top.ticker} ({pct:,.1f}%)"

        web_sourced = sum(1 for h in self.holdings if classify_source(h.yield_source)[1] in
                           ("src_web_primary", "src_web_fallback"))
        web_coverage = f"{web_sourced}/{n}" if n else "0/0"

        fetched_times = [h.last_fetched for h in self.holdings if h.last_fetched]
        last_refreshed = max(fetched_times) if fetched_times else "Never"

        values = {
            "total_weight": f"{total_weight:,.2f}%",
            "cum_ttm_yield": f"{cum_ttm_yield:,.3f}%",
            "cum_sec_yield": f"{cum_sec_yield:,.3f}%",
            "avg_spread": f"{avg_spread:,.1f} bps",
            "hhi": f"{hhi:,.0f}",
            "top_contributor": top_contributor,
            "num_holdings": str(n),
            "web_coverage": web_coverage,
            "last_refreshed": last_refreshed,
            "avg_yield_pair": f"{avg_ttm:,.2f}% / {avg_sec:,.2f}%",
        }
        for key, text in values.items():
            if key in self.stat_cards:
                self.stat_cards[key].configure(text=text)

        # Warn if weights don't sum to ~100%
        if self.holdings and abs(total_weight - 100.0) > 0.01:
            direction = "under" if total_weight < 100 else "over"
            self.weight_warning.configure(
                text=(f"Warning: entered weights sum to {total_weight:,.2f}%, "
                      f"{direction} 100% by {abs(100 - total_weight):,.2f} pts. "
                      f"Cumulative yields above may not reflect the full portfolio.")
            )
        else:
            self.weight_warning.configure(text="")

    def _refresh_pie_chart(self):
        if not MATPLOTLIB_AVAILABLE:
            return
        self.ax.clear()
        pie_colors = [self.ACCENT, self.INFO, self.POSITIVE, "#C77DD1", "#E2685C",
                      "#7DA6D1", "#D1B37D", "#7DD1B0"]
        if self.holdings:
            sorted_holdings = sorted(self.holdings, key=lambda x: x.avg_contribution_pp, reverse=True)
            labels = [h.ticker for h in sorted_holdings]
            sizes = [h.avg_contribution_pp for h in sorted_holdings]

            # Group very small slices into "Other" for readability if many holdings
            if len(sorted_holdings) > 8:
                main = sorted_holdings[:7]
                other_sum = sum(h.avg_contribution_pp for h in sorted_holdings[7:])
                labels = [h.ticker for h in main] + ["Other"]
                sizes = [h.avg_contribution_pp for h in main] + [other_sum]

            wedges, texts, autotexts = self.ax.pie(
                sizes,
                autopct=lambda p: f"{p:.0f}%" if p > 5 else "",
                startangle=90,
                colors=[pie_colors[i % len(pie_colors)] for i in range(len(sizes))],
                textprops={"fontsize": 8, "color": self.BG_APP, "weight": "bold"},
                wedgeprops={"edgecolor": self.BG_PANEL, "linewidth": 1.5},
                pctdistance=0.75,
            )
            self.ax.legend(
                wedges, labels, loc="center left", bbox_to_anchor=(1.0, 0.5),
                fontsize=7, frameon=False, labelcolor=self.TEXT_PRIMARY,
            )
            self.ax.set_title("Share of Dividend Liability", fontsize=9, color=self.TEXT_PRIMARY)
        else:
            self.ax.text(0.5, 0.5, "No holdings yet", ha="center", va="center",
                          fontsize=10, color=self.TEXT_SECONDARY)
            self.ax.axis("off")
        self.fig.patch.set_facecolor(self.BG_PANEL)
        self.ax.set_facecolor(self.BG_PANEL)
        if self.holdings:
            self.fig.subplots_adjust(left=0.02, right=0.60, top=0.88, bottom=0.05)
        else:
            self.fig.tight_layout()
        self.canvas.draw()

    def _refresh_bar_chart(self):
        if not MATPLOTLIB_AVAILABLE:
            return
        self.ax_bar.clear()
        if self.holdings:
            sorted_holdings = sorted(self.holdings, key=lambda x: x.avg_contribution_pp, reverse=True)[:10]
            tickers = [h.ticker for h in reversed(sorted_holdings)]
            ttm_vals = [h.ttm_contribution_pp for h in reversed(sorted_holdings)]
            sec_vals = [h.sec_contribution_pp for h in reversed(sorted_holdings)]

            y = range(len(tickers))
            bar_h = 0.35
            self.ax_bar.barh([p + bar_h / 2 for p in y], ttm_vals, height=bar_h,
                              color=self.INFO, label="TTM contrib. (pp)")
            self.ax_bar.barh([p - bar_h / 2 for p in y], sec_vals, height=bar_h,
                              color=self.ACCENT, label="SEC contrib. (pp)")
            self.ax_bar.set_yticks(list(y))
            self.ax_bar.set_yticklabels(tickers, fontsize=8, color=self.TEXT_PRIMARY)
            self.ax_bar.tick_params(axis="x", labelsize=8, colors=self.TEXT_SECONDARY)
            self.ax_bar.set_title("Top Contributors to Liability (pp)", fontsize=9, color=self.TEXT_PRIMARY)
            self.ax_bar.legend(fontsize=7, facecolor=self.BG_PANEL, edgecolor=self.BORDER,
                                labelcolor=self.TEXT_PRIMARY, loc="lower right")
            self.ax_bar.spines["top"].set_visible(False)
            self.ax_bar.spines["right"].set_visible(False)
            self.ax_bar.spines["left"].set_color(self.BORDER)
            self.ax_bar.spines["bottom"].set_color(self.BORDER)
        else:
            self.ax_bar.text(0.5, 0.5, "No holdings yet", ha="center", va="center",
                              fontsize=10, color=self.TEXT_SECONDARY)
            self.ax_bar.axis("off")
        self.fig_bar.patch.set_facecolor(self.BG_PANEL)
        self.ax_bar.set_facecolor(self.BG_PANEL)
        self.fig_bar.tight_layout()
        self.canvas_bar.draw()


if __name__ == "__main__":
    app = DividendDashboard()
    app.mainloop()
